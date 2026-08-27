#!/usr/bin/env python3
"""MissingFullTextGate — mandatory human-in-the-loop pause for missing PDFs.

Trigger condition (all must hold):
  1. a concrete review topic exists,
  2. API-first literature search has been executed for this run,
  3. at least one paper is screened INCLUDED_PENDING_FULLTEXT or
     HIGH_PRIORITY_PENDING_FULLTEXT,
  4. no legal full-text channel succeeded (Zotero / local / OA / publisher /
     user-provided).

Effects:
  * writes runs/<run-id>/fulltext/missing_fulltext_literature.txt
  * writes runs/<run-id>/fulltext/missing_fulltext_literature.xlsx
  * updates state.json → PAUSED_WAITING_FOR_USER_FULLTEXT
  * downstream stages (synthesis/outline/draft/cite/final) MUST NOT run.

Resume: scripts/resume_helper.py validates newly provided PDFs against the
missing list and clears the gate (or keeps it if high-priority items remain).
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from resume_helper import registry_entry_is_verified

REPO = Path(__file__).resolve().parents[1]

BLOCKING_STATUSES = {"INCLUDED_PENDING_FULLTEXT", "HIGH_PRIORITY_PENDING_FULLTEXT"}
NON_BLOCKING = {
    "EXCLUDED_TITLE_ABSTRACT", "DUPLICATE", "OUT_OF_SCOPE",
    "LOW_PRIORITY_BACKGROUND", "NOT_REQUIRED_FOR_CURRENT_CLAIM",
}

XLSX_FIELDS = [
    "importance_tier", "report_id", "paper_id", "title", "authors", "year",
    "journal_or_source", "doi", "publisher_url", "url", "citation_count",
    "relevance_reason", "screening_status", "fulltext_status",
    "access_attempts", "legal_routes_attempted", "failure_reason",
    "claim_or_section_need", "recommended_user_action", "expected_filename",
    "upload_directory", "user_decision", "provided_local_path",
    "zotero_status", "identity_verified", "resume_command", "notes",
]
TXT_FIELDS = ["importance_tier", "title", "authors", "year", "doi", "url",
              "relevance_reason",
              "fulltext_status", "failure_reason", "recommended_user_action",
              "resume_command"]


def load_screening(run_dir: Path) -> list[dict]:
    """Collect candidates from v3, v2 and legacy screening artifacts."""
    rows: list[dict] = []
    for sc in (run_dir / "screening/adjudicated.csv", run_dir / "screening.csv"):
        if sc.exists():
            with sc.open(encoding="utf-8-sig") as fh:
                rows.extend(dict(r) for r in csv.DictReader(fh))
            break
    reg = run_dir / "literature-registry.jsonl"
    if reg.exists() and not rows:
        for line in reg.read_text(encoding="utf-8").splitlines():
            if line.strip():
                rows.append(json.loads(line))
    literature = run_dir / "literature.json"
    if literature.exists() and not rows:
        rows.extend(json.loads(literature.read_text(encoding="utf-8")))
    registry = run_dir / "fulltext/fulltext-registry.csv"
    by_report: dict[str, dict] = {}
    if registry.exists():
        with registry.open(encoding="utf-8-sig") as fh:
            by_report = {str(r.get("report_id") or ""): dict(r)
                         for r in csv.DictReader(fh)}
    merged = []
    for row in rows:
        report_id = str(row.get("report_id") or row.get("record_id") or "")
        reg = by_report.get(report_id, {})
        merged.append({**row, **{k: v for k, v in reg.items() if v not in (None, "")}})
    return merged


def blocking_rows(rows: list[dict], run_dir: Path | None = None) -> list[dict]:
    out = []
    for r in rows:
        st = (r.get("screening_status") or "").strip()
        ft = (r.get("fulltext_status") or "").strip()
        tier = (r.get("importance_tier") or r.get("priority") or "").strip().lower()
        decision = str(r.get("decision") or "").strip().lower()
        included = (st in BLOCKING_STATUSES or decision in {"include", "included"}
                    or str(r.get("include", "")).lower() in {"true", "1", "yes"})
        identity_ok = str(r.get("identity_verified") or "").strip().casefold() in {
            "true", "1", "yes", "verified", "verified_automatic", "verified_manual",
            "pass", "passed"}
        try:
            page_ok = int(r.get("page_count") or 0) >= 1
        except (TypeError, ValueError):
            page_ok = False
        text_ok = str(r.get("text_quality") or "").strip().casefold() in {
            "acceptable", "good"}
        local_record_ok = bool((r.get("local_path") or r.get("local_pdf") or "").strip()
                               and (r.get("sha256") or "").strip() and identity_ok
                               and (r.get("identity_basis") or "").strip()
                               and (r.get("extracted_text_path") or "").strip()
                               and page_ok and text_ok)
        if run_dir is not None and local_record_ok:
            local_record_ok = registry_entry_is_verified(run_dir, r)
        if (included and (ft not in {
                "AVAILABLE_LOCAL", "AVAILABLE_ZOTERO", "DOWNLOADED_LEGAL"}
                or not local_record_ok)):
            out.append({**r, "_blocking": True})
    # order: HIGH first, then by citation_count desc
    out.sort(key=lambda r: (
        0 if (r.get("importance_tier") or r.get("priority") or "").lower()
        in {"critical", "seminal", "high"} else 1,
        -int(r.get("citation_count") or 0)))
    for i, r in enumerate(out, 1):
        if not r.get("importance_tier"):
            r["importance_tier"] = (
                "critical" if r.get("screening_status") == "HIGH_PRIORITY_PENDING_FULLTEXT"
                else "core")
        if not r.get("recommended_user_action"):
            upload = r.get("upload_directory") or "fulltext/user_uploads"
            r["recommended_user_action"] = (
                f"Place the PDF in {upload} or import it into Zotero.")
        r.setdefault("upload_directory", "fulltext/user_uploads")
        if not r.get("expected_filename"):
            rid = r.get("report_id") or r.get("paper_id") or f"missing-{i:03d}"
            r["expected_filename"] = f"{rid}.pdf"
    return out


def write_txt(rows: list[dict], path: Path) -> None:
    lines = [
        "MISSING FULL-TEXT LITERATURE — ACTION REQUIRED",
        f"generated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}",
        "",
        "These papers passed initial screening as important candidates but no legal",
        "full text could be obtained. The workflow is PAUSED.",
        "Provide PDFs into the run folder or Zotero, then resume;",
        "If an item was incorrectly included, revise and document its exclusion before resuming.",
        "=" * 78, "",
    ]
    for r in rows:
        lines.append(f"[{r.get('importance_tier','?')}] {r.get('title') or '(no title)'}")
        for k in TXT_FIELDS[1:]:
            v = r.get(k)
            if v not in (None, "", []):
                lines.append(f"    {k}: {v}")
        lines.append("-" * 78)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(rows: list[dict], path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Full-text"
    ws.append(XLSX_FIELDS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(f, "") for f in XLSX_FIELDS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(XLSX_FIELDS) + 1):
        field = XLSX_FIELDS[col - 1]
        width = 60 if field == "title" else 42 if field in {
            "url", "publisher_url", "recommended_user_action", "failure_reason"
        } else 28 if field in {"authors", "access_attempts", "legal_routes_attempted",
                               "claim_or_section_need", "provided_local_path"} else 18
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    wb.save(path)
    return True


def update_state(run_dir: Path, extra: dict) -> None:
    sp = run_dir / "state.json"
    state = {}
    if sp.exists():
        try:
            state = json.loads(sp.read_text(encoding="utf-8"))
        except Exception:
            state = {}
    state.update({
        "run_id": run_dir.name,
        **extra,
        "updated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    })
    state.pop("stage", None)
    state["current_stage"] = extra.get("current_stage", "fulltext_acquisition")
    stages = state.setdefault("stages", {})
    if isinstance(stages, dict):
        stages["fulltext"] = (
            "paused" if state.get("status") == "PAUSED_WAITING_FOR_USER_FULLTEXT"
            else "in_progress")
    sp.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    run_dir = Path(args.run_dir)
    if not run_dir.exists():
        print(json.dumps({"error": f"run dir not found: {run_dir}"}))
        return 2

    rows = blocking_rows(load_screening(run_dir), run_dir)
    if not rows:
        update_state(run_dir, {
            "current_stage": "fulltext_acquisition",
            "status": "RUNNING",
            "missing_fulltext_gate": "CLEAR",
        })
        print(json.dumps({"gate": "CLEAR", "message": "No blocking missing full texts."}))
        return 0

    report_dir = run_dir / "fulltext"
    upload_dir = report_dir / "user_uploads"
    report_dir.mkdir(parents=True, exist_ok=True)
    upload_dir.mkdir(parents=True, exist_ok=True)
    for row in rows:
        row["upload_directory"] = str(upload_dir)
        row["resume_command"] = (
            f'python scripts/resume_helper.py validate-pdf --run-dir "{run_dir}"')
        row["recommended_user_action"] = (
            f"Place the identity-matching PDF in {upload_dir} or import it into Zotero. "
            "A scientifically ineligible item may only be removed by a documented screening revision.")
    txt_path = report_dir / "missing_fulltext_literature.txt"
    xlsx_path = report_dir / "missing_fulltext_literature.xlsx"
    xlsx_ok = False
    if not args.dry_run:
        write_txt(rows, txt_path)
        xlsx_ok = write_xlsx(rows, xlsx_path)
        update_state(run_dir, {
            "current_stage": "fulltext_acquisition",
            "status": "PAUSED_WAITING_FOR_USER_FULLTEXT",
            "missing_fulltext_gate": "TRIGGERED",
            "missing_count": len(rows),
            "missing_report_txt": str(txt_path.relative_to(run_dir)),
            "missing_report_xlsx": str(xlsx_path.relative_to(run_dir)) if xlsx_ok else None,
            "fulltext_upload_directory": str(upload_dir),
            "paused_because": (
                "Important included literature lacks legally obtainable full text; "
                "final synthesis/outline/draft/gap-finalization/citation are blocked."),
        })
    print(json.dumps({
        "gate": "TRIGGERED",
        "state": "PAUSED_WAITING_FOR_USER_FULLTEXT",
        "missing_count": len(rows),
        "report_txt": str(txt_path),
        "report_xlsx": str(xlsx_path) if xlsx_ok else None,
        "user_options": [
            "Upload PDFs into the run folder or import into Zotero, then run resume",
            "Revise an inclusion to exclusion only when the protocol shows the item is ineligible",
        ],
        "upload_directory": str(upload_dir),
        "resume_command": f'python scripts/resume_helper.py validate-pdf --run-dir "{run_dir}"',
        "note": "Downstream extraction/synthesis/outline/writing stages are blocked until the gate clears.",
    }, indent=2))
    return 5  # distinctive pause exit code


if __name__ == "__main__":
    sys.exit(main())
